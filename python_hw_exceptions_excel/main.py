from openpyxl import Workbook, load_workbook

def division_calculator():
    try:
        a = float(input("Введите a: "))
        b = float(input("Введите b: "))
        print("Результат:", a / b)

    except ValueError:
        print("Ошибка: введено не число")

    except ZeroDivisionError:
        print("Ошибка: деление на ноль")

    finally:
        print("Завершено")

class Product:
    def __init__(self, title, price, quantity):
        self.title = title
        self.price = price
        self.quantity = quantity

    def line_total(self):
        return self.price * self.quantity


def create_products_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Title", "Price", "Quantity"])

    ws.append([1, "Apple", 10, 3])
    ws.append([2, "Banana", 5, 4])
    ws.append([3, "Orange", "abc", 2])
    ws.append([4, "Pear", 7, "—"])

    wb.save("products.xlsx")


def products_report():
    wb = load_workbook("products.xlsx")
    ws = wb.active

    products = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        _, title, price, quantity = row
        try:
            price = float(price)
            quantity = int(quantity)
            products.append(Product(title, price, quantity))
        except (ValueError, TypeError):
            continue

    report_wb = Workbook()
    report_ws = report_wb.active
    report_ws.append(["Title", "LineTotal"])

    total = 0
    for p in products:
        lt = p.line_total()
        total += lt
        report_ws.append([p.title, lt])

    report_ws.append(["TOTAL", total])
    report_wb.save("products_report.xlsx")

class Student:
    def __init__(self, name):
        self.name = name
        self.scores = []

    def add_score(self, score):
        self.scores.append(score)

    def avg(self):
        return sum(self.scores) / len(self.scores)


def create_students_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Name", "Score1", "Score2", "Score3"])

    ws.append([1, "Ivan", 5, 4, 3])
    ws.append([2, "Anna", 5, "abc", 5])
    ws.append([3, "Oleg", 4, 4, "—"])

    wb.save("students2.xlsx")


def students_report():
    wb = load_workbook("students2.xlsx")
    ws = wb.active

    students = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        _, name, s1, s2, s3 = row
        student = Student(name)

        try:
            for s in (s1, s2, s3):
                student.add_score(float(s))
            students.append(student)
        except (ValueError, TypeError):
            continue

    report_wb = Workbook()
    report_ws = report_wb.active
    report_ws.append(["Name", "Avg"])

    for s in students:
        report_ws.append([s.name, round(s.avg(), 2)])

    report_wb.save("report_students.xlsx")

if __name__ == "__main__":
    create_products_excel()
    create_students_excel()

    division_calculator()
    products_report()
    students_report()
